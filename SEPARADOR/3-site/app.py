import io
import datetime
import unicodedata
import re
from flask import Flask, request, send_file, render_template
from openpyxl import load_workbook, Workbook
from openpyxl.styles import Alignment, Border, Side, Font
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.utils.exceptions import InvalidFileException

# --- IMPORTAÇÕES DE COMPATIBILIDADE ---
# Certifique-se de que pandas, lxml e xlrd estejam instalados
try:
    import pandas as pd
    import lxml
except ImportError:
    pd = None

try:
    import xlrd
    from xlrd import xldate
except ImportError:
    xlrd = None

app = Flask(__name__, template_folder='.', static_folder='.', static_url_path='')

# ==============================================================================
# 1. FUNÇÕES AUXILIARES (Texto, Data e Limpeza)
# ==============================================================================

def normalizar(texto):
    """Remove acentos e deixa maiúsculo."""
    if texto is None: return ""
    return ''.join(c for c in unicodedata.normalize("NFD", str(texto).strip()) 
                   if unicodedata.category(c) != "Mn").upper()

def separar_duas_linhas(texto):
    """Separa texto com quebra de linha (Ex: Nome \n Código)"""
    if not texto: return "", ""
    txt = str(texto).strip()
    if "\n" in txt:
        partes = txt.split("\n")
        return partes[0].strip(), partes[1].strip()
    return txt, ""

def limpar_valor_monetario(valor):
    """Transforma R$ texto em float."""
    if isinstance(valor, (int, float)): return float(valor)
    if isinstance(valor, str):
        v = valor.replace("R$", "").replace(" ", "").replace(".", "").replace(",", ".")
        try: return float(v)
        except: return 0.0
    return 0.0

def converter_data(valor):
    """Tenta converter string para data."""
    if not valor: return None
    if isinstance(valor, (datetime.datetime, datetime.date)): return valor
    
    formatos = ["%d/%m/%Y", "%Y-%m-%d", "%d-%m-%Y", "%Y/%m/%d", "%m-%Y", "%m/%Y"]
    v_str = str(valor).strip().split(' ')[0] # Remove hora se houver
    
    for fmt in formatos:
        try: return datetime.datetime.strptime(v_str, fmt)
        except: continue
    return None

def mes_abreviado(data):
    """Formata data para mmm-aa."""
    meses = {1:"jan", 2:"fev", 3:"mar", 4:"abr", 5:"mai", 6:"jun",
             7:"jul", 8:"ago", 9:"set", 10:"out", 11:"nov", 12:"dez"}
    return f"{meses.get(data.month, '')}-{str(data.year)[-2:]}"

# ==============================================================================
# 2. FUNÇÕES DE LEITURA ESPECÍFICAS (Fallbacks)
# ==============================================================================

def is_html_file(file_stream):
    """Detecta se é um 'fake XLS' (HTML) sem consumir o stream."""
    pos = file_stream.tell()
    file_stream.seek(0)
    try:
        head = file_stream.read(2048)
        # Assinaturas comuns de HTML/XML/UTF-8 BOM
        if b'<html' in head or b'<table' in head or b'\xef\xbb\xbf' in head or b'<?xml' in head:
            return True
    except: pass
    finally:
        file_stream.seek(pos)
    return False

def ler_html_para_workbook(file_stream):
    """Lê HTML via Pandas (simulação Ctrl+C) e retorna um Workbook Openpyxl limpo."""
    if pd is None: raise EnvironmentError("Pandas não instalado para ler HTML.")
    
    file_stream.seek(0)
    # header=None para pegar TUDO como dados, simulando a cópia.
    dfs = pd.read_html(file_stream.read().decode('utf-8', errors='ignore'), header=None, decimal=',', thousands='.')
    
    if not dfs: raise ValueError("Nenhuma tabela encontrada no HTML.")
    
    df = max(dfs, key=lambda x: x.size)
    wb_limpo = Workbook()
    ws_limpo = wb_limpo.active

    for row in dataframe_to_rows(df, index=False, header=False):
        row_limpa = ["" if pd.isna(x) else x for x in row]
        ws_limpo.append(row_limpa)
        
    return wb_limpo

def ler_xls_binario_para_workbook(file_stream):
    """Lê XLS binário via xlrd e retorna um Workbook Openpyxl limpo."""
    if xlrd is None: raise EnvironmentError("xlrd não instalado para ler XLS binário.")
    
    file_stream.seek(0)
    book = xlrd.open_workbook(file_contents=file_stream.read(), formatting_info=False)
    sheet = book.sheet_by_index(0)
    
    wb_limpo = Workbook()
    ws_limpo = wb_limpo.active

    for r in range(sheet.nrows):
        row_vals = []
        for c in range(sheet.ncols):
            val = sheet.cell_value(r, c)
            if sheet.cell_type(r, c) == xlrd.XL_CELL_DATE:
                try: val = xldate.xldate_as_datetime(val, book.datemode)
                except: pass
            row_vals.append(val)
        ws_limpo.append(row_vals)
        
    return wb_limpo

# ==============================================================================
# 3. FUNÇÃO CENTRAL DE CARREGAMENTO (FLUXO ROBUSTO DE PRIORIDADE)
# ==============================================================================

def carregar_workbook_inicial(file_stream_bytes, filename):
    """
    Tenta carregar o arquivo na ordem correta:
    1. XLSX Padrão (OpenPyXL)
    2. HTML/XML (Fake XLS via Pandas)
    3. XLS Binário (xlrd)
    """
    
    # 1. TENTA CARREGAR COMO FORMATO OPENPYXL PADRÃO (.xlsx, .xlsm, etc.)
    try:
        print("LOG: [PRIORIDADE] Tentando carregar como XLSX/XLSM padrão...")
        stream_xlsx = io.BytesIO(file_stream_bytes)
        return load_workbook(stream_xlsx)
    except InvalidFileException:
        print("LOG: Falhou no XLSX. O arquivo não é um formato OpenPyXL válido.")
        pass
    except Exception as e:
        print(f"LOG: Erro inesperado ao carregar XLSX/XLSM, tentando fallback: {e}")
        pass

    # 2. TENTA LER COMO HTML (FAKE XLS)
    stream_html_detect = io.BytesIO(file_stream_bytes)
    if is_html_file(stream_html_detect):
        print("LOG: Detectado HTML (Fake XLS). Lendo com Pandas...")
        stream_html_read = io.BytesIO(file_stream_bytes)
        return ler_html_para_workbook(stream_html_read)
    
    # 3. TENTA LER COMO XLS BINÁRIO ANTIGO
    elif filename.endswith('.xls'):
        stream_xls_read = io.BytesIO(file_stream_bytes)
        try:
            print("LOG: Detectado XLS Binário. Lendo com xlrd...")
            return ler_xls_binario_para_workbook(stream_xls_read)
        except Exception as e:
            # Caso o xlrd falhe (porque o XLS é na verdade HTML malfeito)
            print(f"LOG: Falha no xlrd. Tentativa final via HTML/Pandas: {e}")
            stream_html_read = io.BytesIO(file_stream_bytes)
            return ler_html_para_workbook(stream_html_read)

    raise ValueError("Formato de arquivo não suportado, corrompido ou incompatível.")

# ==============================================================================
# 4. LÓGICA DE FILTRO E PROCESSAMENTO
# ==============================================================================

def encontrar_cabecalho(ws):
    """Procura em qual linha está o cabeçalho real (LANÇAMENTO/VALOR)"""
    for r in range(1, 20):
        row_vals = [normalizar(c.value) for c in ws[r]]
        txt_row = " ".join(row_vals)
        if "LANCAMENTO" in txt_row and "VALOR" in txt_row:
            return r
    return 1 # Assume linha 1 como fallback

def processar_arquivo_limpo(wb_entrada, meses, ano):
    ws_entrada = wb_entrada.active
    
    linha_header = encontrar_cabecalho(ws_entrada)
    
    # Mapeamento de Colunas (usando a linha de cabeçalho encontrada)
    mapa = {}
    for cell in ws_entrada[linha_header]:
        if cell.value:
            mapa[normalizar(cell.value)] = cell.column

    # Mapeamento de índices (Baseado na estrutura da sua planilha)
    idx_lanc = mapa.get("LANCAMENTO") or 1
    idx_rubrica = mapa.get("RUBRICA") or 3
    idx_tipo = mapa.get("TIPO DOCUMENTO") or mapa.get("TIPO DOCUM") or 4
    idx_comp = mapa.get("COMPETENCIA") or mapa.get("COMPETENC") or 5
    idx_data = mapa.get("DATA PAGAMENTO") or mapa.get("DATA PAGAM") or 6
    idx_valor = mapa.get("VALOR") or 7
    idx_sit = mapa.get("SITUACAO") or 8

    # Cria arquivo final
    wb_final = Workbook()
    ws_final = wb_final.active
    ws_final.append(["CÓDIGO", "FORNECEDOR", "RUBRICA", "DOCUMENTO", "COMPETÊNCIA", "DATA PAGAMENTO", "SITUAÇÃO", "VALOR"])

    for row in ws_entrada.iter_rows(min_row=linha_header + 1, values_only=True):
        if not any(row): continue
        
        try:
            raw_comp = row[idx_comp - 1]
            dt_comp = converter_data(raw_comp)
        except: continue

        # FILTRO DE MÊS/ANO (AGORA COM LISTA DE MESES)
        if dt_comp and dt_comp.month in meses and dt_comp.year == ano:
            
            # --- CORREÇÃO DE EXTRAÇÃO DE CÓDIGO/FORNECEDOR ---
            raw_lanc = row[idx_lanc - 1] if idx_lanc <= len(row) else ""
            full_lanc_text = str(raw_lanc).strip()
            
            codigo = ""
            forn = full_lanc_text

            if "\n" in full_lanc_text:
                # Caso 1: Formato ideal (Nome \n Código)
                partes = full_lanc_text.split("\n", 1)
                forn = partes[0].strip()
                cod_txt = partes[1].strip()
                codigo = cod_txt.replace("Cod.:", "").replace("Cod:", "").strip()
            elif "COD:" in normalizar(full_lanc_text):
                # Caso 2: Linha única contendo apenas o código (Ex: "Cod.: 12345")
                codigo = full_lanc_text.replace("Cod.:", "").replace("Cod:", "").strip()
                forn = "" # O fornecedor está vazio nesta célula
            # Se não houver '\n' nem "COD:", o código permanece vazio e 'forn' contém o texto completo.
            
            # 2. TIPO DOCUMENTO (Tipo e Doc) -> Mantém a lógica original
            raw_tipo = row[idx_tipo - 1] if idx_tipo <= len(row) else ""
            tipo, doc_txt = separar_duas_linhas(raw_tipo)
            documento = doc_txt.replace("Doc.:", "").replace("Doc:", "").strip()

            # 3. Campos Diretos e Valor
            rubrica = row[idx_rubrica - 1] if idx_rubrica <= len(row) else ""
            situacao = row[idx_sit - 1] if idx_sit <= len(row) else ""
            raw_valor = row[idx_valor - 1] if idx_valor <= len(row) else 0
            val_float = limpar_valor_monetario(raw_valor)
            dt_pag = converter_data(row[idx_data - 1]) if idx_data <= len(row) else None

            ws_final.append([codigo, forn, rubrica, documento, dt_comp, dt_pag, situacao, val_float])

    # Formatação
    thin = Side(style="thin")
    borda = Border(left=thin, right=thin, top=thin, bottom=thin)
    
    for row in ws_final.iter_rows():
        for cell in row:
            cell.border = borda
            if cell.row == 1:
                cell.font = Font(bold=True)
                cell.alignment = Alignment(horizontal='center')
            else:
                if cell.column in [1, 5, 6]: cell.alignment = Alignment(horizontal='center')
                elif cell.column == 8: 
                    cell.alignment = Alignment(horizontal='right')
                    cell.number_format = '#,##0.00'
                
                if cell.column == 5 and isinstance(cell.value, datetime.datetime):
                    cell.value = mes_abreviado(cell.value)
                if cell.column == 6 and isinstance(cell.value, datetime.datetime):
                    cell.number_format = 'dd/mm/yyyy'

    out = io.BytesIO()
    wb_final.save(out)
    out.seek(0)
    return out

# ==============================================================================
# 5. ROTAS FLASK
# ==============================================================================

@app.route("/")
def index():
    return render_template("site.html")

@app.route("/processar", methods=["POST"])
def processar():
    if 'file' not in request.files: return "Sem arquivo", 400
    file = request.files['file']
    if not file.filename: return "Vazio", 400

    try:
        # CAPTURA DE DADOS: Lista de meses e nome personalizado
        meses = [int(x) for x in request.form.getlist('meses')]
        ano = int(request.form.get('ano'))
        nome_arquivo_user = request.form.get('nome_arquivo')
        
        # Lê o conteúdo binário do arquivo na memória APENAS UMA VEZ
        file_bytes = file.read()
        filename = file.filename.lower()
        
        # 1. Carrega o arquivo usando a função robusta
        wb_entrada = carregar_workbook_inicial(file_bytes, filename)
        
        # 2. Processamento e filtragem (passando a lista de meses)
        print("LOG: Processando dados e filtrando...")
        output = processar_arquivo_limpo(wb_entrada, meses, ano)
        
        # 3. Definição do Nome do Arquivo de Saída
        if nome_arquivo_user and nome_arquivo_user.strip():
            nome_final = nome_arquivo_user.strip()
            # Garante que tenha a extensão .xlsx
            if not nome_final.lower().endswith('.xlsx'):
                nome_final += '.xlsx'
        else:
            # Nome padrão se o usuário não preencher
            lista_meses_str = "_".join(map(str, meses))
            nome_final = f"processado_meses_{lista_meses_str}_{ano}.xlsx"

        return send_file(output, as_attachment=True, download_name=nome_final, mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')

    except Exception as e:
        error_message = f"Não foi possível processar o arquivo. Detalhe: {str(e)}"
        
        if "Nenhuma tabela encontrada" in str(e):
            error_message = "Não foi possível extrair a tabela do arquivo (Verifique se é um XLS/HTML válido)."
        if "Formato de arquivo não suportado" in str(e):
            error_message = "Formato de arquivo não suportado ou arquivo corrompido."
        if "No such file or directory" in str(e):
            error_message = "Erro de leitura do arquivo. Certifique-se de que o arquivo não está sendo usado por outro programa."
        
        print(f"ERRO CRÍTICO: {e}")
        return f"Erro: {error_message}", 500

if __name__ == "__main__":
    app.run(debug=True)
